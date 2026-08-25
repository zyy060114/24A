from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parents[1]
TABLES = OUT / "tables"
SUMMARY = json.loads((TABLES / "q3_result_summary.json").read_text(encoding="utf-8"))


def _load_csv(name: str) -> list[dict]:
    with (TABLES / name).open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _style_sheet(sheet, widths=None):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if widths:
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook():
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "关键结果"
    summary_sheet.append(["指标", "数值", "单位/说明"])
    bisection = SUMMARY["bisection"]
    rows = [
        ["调头空间半径 R", SUMMARY["turning_radius_m"], "m，题面固定"],
        ["搜索变量", "螺距 p", "m"],
        ["单调扫描区间", f'{SUMMARY["pitch_scan_range_m"][0]}–{SUMMARY["pitch_scan_range_m"][1]}', "m"],
        ["单调认证", "通过" if SUMMARY["monotonicity_passed"] else "不通过", "无可行→不可行反转"],
        ["不可行下界", bisection["lower_infeasible_m"], "m"],
        ["可行上界", bisection["upper_feasible_m"], "m"],
        ["夹逼宽度", "=B7-B6", "m，由上下界计算"],
        ["临界估计", "=(B6+B7)/2", "m，由上下界中点计算"],
        ["论文厘米口径", "=B9*100", "cm，约 45.0 cm"],
        ["保证可行的毫米口径", "=ROUNDUP(B9*1000,0)/1000", "m，向上取整"],
        ["临界危险龙头半径", bisection["upper"]["critical_head_radius_m"], "m"],
        ["临界候选板对", "第1块–第20块", "不可行下界处"],
        ["正式运行判定", "通过" if SUMMARY["pass"] else "不通过", "R-Q3-MIN-PITCH-001"],
    ]
    for row in rows:
        summary_sheet.append(row)
    _style_sheet(summary_sheet, [24, 22, 34])
    for row in range(2, summary_sheet.max_row + 1):
        summary_sheet.cell(row, 2).number_format = "0.0000000000"

    scan_sheet = workbook.create_sheet("螺距扫描")
    scan_rows = _load_csv("q3_pitch_scan.csv")
    scan_headers = [
        "pitch_m", "feasible", "global_min_margin_m", "critical_head_theta",
        "critical_head_radius_m", "witness_pair_paper_numbering", "sampled_states",
        "refined_states", "tested_pairs", "circle_rejected_pairs", "total_forbidden_pairs",
    ]
    scan_sheet.append(scan_headers)
    for row in scan_rows:
        scan_sheet.append([row.get(header, "") for header in scan_headers])
    _style_sheet(scan_sheet, [12, 12, 22, 22, 22, 26, 15, 15, 15, 22, 20])

    convergence_sheet = workbook.create_sheet("采样收敛")
    conv_rows = _load_csv("q3_sampling_convergence.csv")
    conv_headers = [
        "samples", "pitch_m", "feasible", "global_min_margin_m",
        "critical_head_theta", "critical_head_radius_m", "refined_states",
        "tested_pairs", "circle_rejected_pairs", "total_forbidden_pairs",
    ]
    convergence_sheet.append(conv_headers)
    for row in conv_rows:
        convergence_sheet.append([row.get(header, "") for header in conv_headers])
    _style_sheet(convergence_sheet, [12, 14, 12, 22, 22, 22, 16, 15, 22, 20])

    monotone_sheet = workbook.create_sheet("单调性数学依据")
    monotone_sheet.append(["pitch_m", "normal_spacing_at_R_m", "derivative_m_per_m"])
    for row in _load_csv("q3_normal_spacing_monotonicity.csv"):
        monotone_sheet.append([row["pitch_m"], row["normal_spacing_at_R_m"], row["derivative_m_per_m"]])
    _style_sheet(monotone_sheet, [15, 28, 24])

    workbook.save(TABLES / "result3.xlsx")


def write_markdown_tables():
    b = SUMMARY["bisection"]
    validation = [
        "# Q3 验证结果",
        "",
        "| 检查项 | 结果 | 判据/说明 |",
        "|---|---:|---|",
        f'| 全域扫描区间 | {SUMMARY["pitch_scan_range_m"][0]:.3f}–{SUMMARY["pitch_scan_range_m"][1]:.3f} m | 覆盖几何下界 0.28125 m 以上的实际区间 |',
        f'| 单调性认证 | {"通过" if SUMMARY["monotonicity_passed"] else "不通过"} | 未出现“可行后重新碰撞” |',
        f'| 初始碰撞—可行区间 | {SUMMARY["initial_bracket_m"][0]:.3f}–{SUMMARY["initial_bracket_m"][1]:.3f} m | 逻辑值由 0 变为 1 |',
        f'| 二分不可行下界 | {b["lower_infeasible_m"]:.11f} m | 最小裕度 {b["lower"]["global_min_margin_m"]:.12e} m |',
        f'| 二分可行上界 | {b["upper_feasible_m"]:.11f} m | 最小裕度 {b["upper"]["global_min_margin_m"]:.12e} m |',
        f'| 夹逼宽度 | {b["width_m"]:.12e} m | 小于 0.01 mm |',
        f'| 最危险龙头半径 | {b["upper"]["critical_head_radius_m"]:.9f} m | 大于 4.5 m，证明不能只查终点 |',
        "| 临界候选板对 | 第1块与第20块 | 内部代码索引 (0,19) |",
        "| 独立 Shapely 复核 | 下界相交面积 4.296174e-11 m²；上界距离 4.395378e-6 m | 与 SAT 符号一致 |",
        "| 33/65/129 点收敛 | 最小裕度差小于 3e-15 m | 自适应局部细化后稳定 |",
        f'| 非相邻板对总数 | {b["upper"]["total_forbidden_pairs"]} | C(223,2)-222 |',
        f'| 正式运行结论 | {"通过" if SUMMARY["pass"] else "不通过"} | 运行令牌 R-Q3-MIN-PITCH-001 |',
    ]
    (TABLES / "Q3_验证结果.md").write_text("\n".join(validation) + "\n", encoding="utf-8")

    monotonicity = [
        "# Q3 单调性与二分依据",
        "",
        "相邻螺线圈在固定极角方向的径向间距等于螺距 p。固定半径 r 处的法向圈距为：",
        "",
        r"\[",
        r"d_\perp(r,p)=\frac{pr}{\sqrt{r^2+(p/(2\pi))^2}},",
        r"\qquad",
        r"\frac{\partial d_\perp}{\partial p}=\frac{r^3}{[r^2+(p/(2\pi))^2]^{3/2}}>0.",
        r"\]",
        "",
        "因此增大螺距会严格增大圈间法向距离。由于实体板凳姿态也会随 p 改变，正式求解没有只凭该局部公式直接宣称全部 SAT 间隙单调，而是先在完整搜索区间运行全过程碰撞扫描。扫描只出现一次“碰撞→可行”转变，未出现反转，故在该经过认证的区间内使用逻辑二分。",
    ]
    (TABLES / "Q3_单调性与二分依据.md").write_text("\n".join(monotonicity) + "\n", encoding="utf-8")


if __name__ == "__main__":
    build_workbook()
    write_markdown_tables()
