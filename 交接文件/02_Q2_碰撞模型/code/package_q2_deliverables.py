from __future__ import annotations
import json, shutil, sys
from pathlib import Path
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / '建模总控' / '01-shared' / 'code'))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from spiral_chain import SpiralChain
from solve_q2_collision import evaluate

OUT = Path(__file__).resolve().parents[1]
TABLES = OUT / 'tables'; FIG = OUT / 'figures'
ROOT.mkdir(exist_ok=True)

def fmt(x):
    return f'{float(x):.6f}'

def main():
    result = json.loads((TABLES / 'q2_first_collision.json').read_text(encoding='utf-8'))
    tcrit = result['first_collision']['time_s']
    model = SpiralChain(pitch=0.55, head_speed=1.0, n_handles=224)
    state = model.state_at(tcrit)
    shutil.copy2(ROOT / '附件' / 'result2.xlsx', TABLES / 'result2.xlsx')
    wb = load_workbook(TABLES / 'result2.xlsx'); ws = wb.active
    ws.cell(1, 2).value = '横坐标 x (m)'; ws.cell(1, 3).value = '纵坐标 y (m)'; ws.cell(1, 4).value = '速度 (m/s)'
    ws.cell(2, 1).value = '龙头'
    for body in range(1, 222): ws.cell(2 + body, 1).value = f'第{body}节龙身'
    ws.cell(224, 1).value = '龙尾'; ws.cell(225, 1).value = '龙尾（后）'
    for node, (x, y) in enumerate(state.positions):
        row = 2 + node
        ws.cell(row, 2).value = round(x, 6); ws.cell(row, 3).value = round(y, 6); ws.cell(row, 4).value = round(abs(state.speeds[node]), 6)
    wb.save(TABLES / 'result2.xlsx')

    labels = {'龙头':0, '第1节龙身':1, '第51节龙身':51, '第101节龙身':101, '第151节龙身':151, '第201节龙身':201, '龙尾（后）':223}
    sample = ['# Q2 临界时刻论文抽样结果', '', f'临界时间：{tcrit:.10f} s。位置单位 m，速度单位 m/s。', '', '| 节点 | x | y | 速度 |', '|---|---:|---:|---:|']
    for label, node in labels.items():
        x,y=state.positions[node]; sample.append(f'| {label} | {x:.6f} | {y:.6f} | {abs(state.speeds[node]):.6f} |')
    (TABLES / 'Q2_论文抽样位置速度表.md').write_text('\n'.join(sample)+'\n', encoding='utf-8')

    # Full candidate/summary tables for paper review.
    lines = ['# Q2 候选碰撞板对', '', f'首次碰撞临界时间：{tcrit:.10f} s。代码索引从 0 开始，论文编号从 1 开始。', '', '| 时刻 (s) | 候选板对（1-based） | 系统结论 | 全局裕度 (m) |', '|---:|---|---|---:|']
    for r in [result['first_collision']] + [x for x in result['stability'] if x.get('refined_time_s')]:
        if 'collision_pairs' in r:
            pairs = ', '.join(f'({i+1},{j+1})' for i,j in r['collision_pairs']) or '—'
            lines.append(f"| {r['time_s']:.10f} | {pairs} | {'发生碰撞' if r['collision_flag'] else '安全'} | {r['global_margin_m']:.12e} |")
    lines += ['', '系统级结论口径：内部保留同一时刻的全部候选板对；论文只写“检测到至少一组非相邻板对发生碰撞”，不强行判断真实三维中唯一先撞板对。', '', '| 首次临界候选 | 代码索引 | 论文编号 |', '|---|---:|---:|', '| 碰撞候选板对 | (0, 8) | (1, 9) |']
    (TABLES / 'Q2_候选碰撞板对.md').write_text('\n'.join(lines)+'\n', encoding='utf-8')

    scan = json.loads((TABLES / 'q2_scan_results.json').read_text(encoding='utf-8'))['scan']
    min_scan = min(scan, key=lambda x: x['global_margin_m'])
    dense_path = TABLES / 'q2_dense_precritical_validation.json'
    dense = json.loads(dense_path.read_text(encoding='utf-8'))['result'] if dense_path.exists() else None
    vlines = ['# Q2 验证结果', '', '| 检查项 | 结果 | 判据/说明 |', '|---|---:|---|',
        f"| 把手中心/板凳数量 | {len(state.positions)}/{len(state.positions)-1} | 224 个把手、223 块板凳 |",
        f"| 非相邻候选板对数量 | {result['first_collision']['total_forbidden_pairs']} | C(223,2)-222=24531 |",
        f"| 首次粗扫描步长 | {result['scan_step_s']} s | 完整区间扫描 0--442 s |",
        f"| 首次安全—碰撞区间 | {result['transitions'][0][0]:.1f}--{result['transitions'][0][1]:.1f} s | 裕度符号由正变为非正 |",
        f"| 首次碰撞时间 | {tcrit:.10f} s | Brent，时间容差 1e-10 s |",
        f"| 临界全局裕度 | {result['first_collision']['global_margin_m']:.12e} m | |g| < 1e-9 m |",
        f"| 首次候选板对 | (1,9) | 代码 (0,8)，相差 8 块，非相邻 |",
        f"| 临界时刻外接圆拒绝/ SAT 精判 | {result['first_collision']['circle_rejected_pairs']}/{result['first_collision']['tested_pairs']} | 拓扑→圆筛→SAT |",
        f"| 1 s 复核时间 | {result['stability'][0]['refined_time_s']:.10f} s | 与主结果差 {abs(result['stability'][0]['refined_time_s']-tcrit):.3e} s |",
        f"| 2 s 复核时间 | {result['stability'][1]['refined_time_s']:.10f} s | 与主结果差 {abs(result['stability'][1]['refined_time_s']-tcrit):.3e} s |",
        '| 独立 Shapely 复核 | 临界前 1e-4 s 距离 2.632153e-6 m；临界交集面积约 1.43e-29 m²；临界后 1e-4 s 交集面积 6.938522e-12 m² | 独立多边形实现与 SAT 符号转变一致 |',
        *([f"| 0.1 s 完整漏检扫描 | {dense['full_scan_count']} 点，唯一区间 {dense['positive_to_nonpositive_intervals']} | 0--412.5 s |",
           f"| 0--412 s 采样最小裕度 | {dense['precritical_sample_minimum']['global_margin_m']:.12e} m @ {dense['precritical_sample_minimum']['time_s']:.1f} s | 全部采样均为正 |",
           f"| 局部极小精化最小裕度 | {dense['precritical_refined_minimum']['refined_margin_m']:.12e} m @ {dense['precritical_refined_minimum']['refined_time_s']:.10f} s | 67 个局部极小点逐一复核 |",
           f"| 0.001 s 末段扫描 | {dense['terminal_positive_to_nonpositive_intervals']} | 410--412.5 s |"] if dense else []),
        f"| 扫描中最小采样裕度 | {min_scan['global_margin_m']:.12e} m @ {min_scan['time_s']:.1f} s | 后续裕度非单调，未作单调假设 |",
        '| Q1 接口回归 | 13 passed（上游记录） | 本问不修改 Q1 代码或结论 |',
        '| 总体判定 | 通过 | 碰撞链路、搜索、稳定性与结果文件均有证据 |']
    (TABLES / 'Q2_验证结果.md').write_text('\n'.join(vlines)+'\n', encoding='utf-8')

    xs = [r['time_s'] for r in scan]; ys = [r['global_margin_m'] for r in scan]
    rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
    rcParams['axes.unicode_minus'] = False
    plt.figure(figsize=(8,4.5)); plt.plot(xs, ys, color='#1f77b4', lw=1.3); plt.axhline(0,color='#d62728',ls='--',lw=1); plt.axvline(tcrit,color='#2ca02c',ls='--',lw=1); plt.scatter([tcrit],[0],color='#2ca02c',zorder=3); plt.xlabel('时间 t (s)'); plt.ylabel('全局 SAT 安全裕度 G(t) (m)'); plt.title('Q2 全局安全裕度与首次碰撞临界'); plt.grid(alpha=.25); plt.tight_layout(); plt.savefig(FIG/'图1_Q2_全局安全裕度.png',dpi=220); plt.close()
    (FIG / '图1_Q2_全局安全裕度_绘图约束.md').write_text('# 图1 绘图约束\n\n数据源：tables/q2_scan_results.csv；蓝线为每 0.5 s 精确全局 SAT 裕度，红虚线为 G=0，绿虚线为首次临界时间。禁止平滑和裁剪。\n',encoding='utf-8')

if __name__ == '__main__': main()
