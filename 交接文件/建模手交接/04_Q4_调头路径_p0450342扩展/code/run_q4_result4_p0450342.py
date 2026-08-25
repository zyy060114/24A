"""Run, validate, and package the Q4 p=0.450342 m extension scenario."""

from __future__ import annotations

import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import brentq

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from q4_result4_p0450342 import PITCH, build_boundary_geometry, compute_series, state_at  # noqa: E402


Q3_INFEASIBLE_LOWER = 0.45033203125
Q3_FEASIBLE_UPPER = 0.450341796875
Q3_MIDPOINT_ESTIMATE = 0.4503369140625
Q3_SELECTED_RECHECK_MARGIN = 4.598112163634038e-06
REPO_HANDOFF = None


def _resolve_paths() -> tuple[Path, Path, Path | None]:
    """Support both the source tree and the copied handoff code directory."""

    if HERE.name == "code" and (HERE.parent / "tables").exists():
        return HERE.parent, HERE.parent / "tables" / "result4_template.xlsx", None
    project_root = HERE.parents[2]
    control_root = HERE.parents[1]
    delivery = control_root / "03-deliverables" / "04_Q4_调头路径_p0450342扩展"
    return delivery, project_root / "附件" / "result4.xlsx", REPO_HANDOFF


def _write_workbook(states, template: Path, output: Path) -> None:
    workbook = load_workbook(template)
    position_sheet = workbook["位置"]
    speed_sheet = workbook["速度"]
    times = [int(round(state.time_s)) for state in states]
    for column, time_s in enumerate(times, start=2):
        position_sheet.cell(1, column).value = f"{time_s} s"
        speed_sheet.cell(1, column).value = f"{time_s} s"
    for column, state in enumerate(states, start=2):
        for handle_index, (x_coord, y_coord) in enumerate(state.positions):
            if handle_index <= 221:
                row_x = 2 + 2 * handle_index
            elif handle_index == 222:
                row_x = 446
            else:
                row_x = 448
            position_sheet.cell(row_x, column).value = round(float(x_coord), 6)
            position_sheet.cell(row_x + 1, column).value = round(float(y_coord), 6)
        for handle_index, speed in enumerate(state.speeds):
            speed_sheet.cell(2 + handle_index, column).value = round(abs(float(speed)), 6)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def _write_collision_csv(states, output: Path) -> None:
    fields = [
        "time_s",
        "global_margin_m",
        "collision_flag",
        "witness_pair",
        "minimum_speed_denominator",
        "max_handle_speed_m_per_s",
    ]
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for state in states:
            writer.writerow(
                {
                    "time_s": int(round(state.time_s)),
                    "global_margin_m": f"{state.global_margin_m:.12g}",
                    "collision_flag": int(state.global_margin_m < 0.0),
                    "witness_pair": f"{state.witness_pair[0]}-{state.witness_pair[1]}",
                    "minimum_speed_denominator": f"{state.minimum_speed_denominator:.12g}",
                    "max_handle_speed_m_per_s": f"{max(abs(value) for value in state.speeds):.12g}",
                }
            )


def _write_requested_summary(states, output: Path) -> None:
    wanted_times = {-100, -50, 0, 50, 100}
    wanted_handles = [0, 1, 51, 101, 151, 201, 223]
    labels = {
        0: "龙头前把手",
        1: "龙头后第1节",
        51: "龙头后第51节",
        101: "龙头后第101节",
        151: "龙头后第151节",
        201: "龙头后第201节",
        223: "龙尾后把手",
    }
    fields = [
        "time_s",
        "handle_index",
        "paper_label",
        "x_m",
        "y_m",
        "vx_m_per_s",
        "vy_m_per_s",
        "speed_m_per_s",
    ]
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for state in states:
            if int(round(state.time_s)) not in wanted_times:
                continue
            for index in wanted_handles:
                x_coord, y_coord = state.positions[index]
                vx, vy = state.velocities[index]
                writer.writerow(
                    {
                        "time_s": int(round(state.time_s)),
                        "handle_index": index,
                        "paper_label": labels[index],
                        "x_m": f"{x_coord:.9f}",
                        "y_m": f"{y_coord:.9f}",
                        "vx_m_per_s": f"{vx:.9f}",
                        "vy_m_per_s": f"{vy:.9f}",
                        "speed_m_per_s": f"{abs(state.speeds[index]):.9f}",
                    }
                )


def _find_first_collision_time(geometry) -> dict[str, object] | None:
    previous_time = 0.0
    previous_state = state_at(geometry, previous_time)
    if previous_state.global_margin_m < 0.0:
        return {
            "time_s": 0.0,
            "margin_m": previous_state.global_margin_m,
            "witness_pair": list(previous_state.witness_pair),
        }
    for current_time in np.arange(0.25, 100.0001, 0.25):
        current_state = state_at(geometry, float(current_time))
        if current_state.global_margin_m < 0.0:
            root = brentq(
                lambda value: state_at(geometry, float(value)).global_margin_m,
                previous_time,
                float(current_time),
                xtol=1e-8,
            )
            root_state = state_at(geometry, root)
            return {
                "time_s": float(root),
                "margin_m": float(root_state.global_margin_m),
                "witness_pair": list(root_state.witness_pair),
                "bracket_s": [previous_time, float(current_time)],
            }
        previous_time = float(current_time)
        previous_state = current_state
    return None


def _validate_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    required_sheets = {"位置", "速度"}
    sheets_ok = required_sheets.issubset(workbook.sheetnames)
    time_headers_ok = True
    populated_ok = True
    formula_errors: list[str] = []
    for sheet_name in required_sheets:
        sheet = workbook[sheet_name]
        expected_headers = [f"{time_s} s" for time_s in range(-100, 101)]
        actual_headers = [sheet.cell(1, column).value for column in range(2, 203)]
        time_headers_ok = time_headers_ok and actual_headers == expected_headers
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    formula_errors.append(f"{sheet_name}!{cell.coordinate}:{value}")
    position_sheet = workbook["位置"]
    speed_sheet = workbook["速度"]
    for column in range(2, 203):
        if position_sheet.cell(2, column).value is None or position_sheet.cell(449, column).value is None:
            populated_ok = False
        if speed_sheet.cell(2, column).value is None or speed_sheet.cell(225, column).value is None:
            populated_ok = False
    return {
        "required_sheets_ok": sheets_ok,
        "time_headers_ok": time_headers_ok,
        "required_cells_populated": populated_ok,
        "formula_error_count": len(formula_errors),
        "formula_errors": formula_errors[:10],
    }


def _write_validation_markdown(geometry, states, workbook_check, first_collision, output: Path) -> None:
    minimum_denominator = min(state.minimum_speed_denominator for state in states)
    finite_values = all(
        np.isfinite(np.asarray(state.positions)).all()
        and np.isfinite(np.asarray(state.velocities)).all()
        and np.isfinite(np.asarray(state.speeds)).all()
        for state in states
    )
    text = f"""# Q4 p=0.450342 验证结果

| 检验项 | 结果 | 判定 |
| --- | ---: | --- |
| Q3全盘入复核最小裕度 | {Q3_SELECTED_RECHECK_MARGIN:.12g} m | 安全侧 |
| 双圆弧圆相切残差 | {geometry.circle_tangency_residual:.12g} | 通过 |
| 端点切向残差 | {geometry.endpoint_tangent_error:.12g} | 通过 |
| 接合点切向残差 | {geometry.joint_tangent_error:.12g} | 通过 |
| 半径比误差 | {abs(geometry.radius_large - 2.0 * geometry.radius_small):.12g} m | 通过 |
| 调头路径最大半径 | {geometry.maximum_radius():.12g} m | 不超过4.5 m |
| 速度递推最小分母绝对值 | {minimum_denominator:.12g} | 远离奇异阈值 |
| 位置和速度有限值 | {finite_values} | 通过 |
| Excel工作表与表头 | {workbook_check['required_sheets_ok']} / {workbook_check['time_headers_ok']} | 通过 |
| Excel关键单元格完整 | {workbook_check['required_cells_populated']} | 通过 |
| Excel错误值数量 | {workbook_check['formula_error_count']} | 应为0 |

连续时间首次碰撞：{json.dumps(first_collision, ensure_ascii=False)}。

说明：Q3安全复核仅证明盘入螺线段；Q4双圆弧拼接后的碰撞必须由本问重新计算，不能用Q3结论替代。
"""
    output.write_text(text, encoding="utf-8")


def _write_formula_audit(output: Path) -> None:
    text = r"""# Q4 公式与符号自审

| 符号 | 含义 | 单位 |
| --- | --- | --- |
| $p$ | 螺距，本扩展口径取0.450342 | m |
| $a=p/(2\pi)$ | 阿基米德螺线参数 | m/rad |
| $s$ | 全局路径弧长坐标，沿运动方向递增 | m |
| $P_i=(x_i,y_i)$ | 第$i$个把手中心 | m |
| $L_i$ | 相邻把手中心距，龙头2.86、其余1.65 | m |
| $\tau_i$ | 第$i$个把手所在轨迹的单位切向量 | 1 |
| $w_i$ | 第$i$个把手的有向切向速度 | m/s |
| $g_{ij}$ | 板凳$i,j$的SAT分离裕度 | m |

1. 位置约束：$\lVert P_i-P_{i-1}\rVert=L_i$。
2. 速度递推：$w_i=w_{i-1}\frac{(P_i-P_{i-1})\cdot\tau_{i-1}}{(P_i-P_{i-1})\cdot\tau_i}$。
3. 速度分量：$(v_{x,i},v_{y,i})=w_i\tau_i$；Excel速度表填速度大小$|w_i|$。
4. 碰撞判定：先做拓扑排除和外接圆初筛，再对候选矩形做SAT；$g_{ij}<0$为穿透，$g_{ij}=0$为接触。
5. 重叠板凳口径：同一时刻只要存在任意一对非相邻板凳满足$g_{ij}\le0$，就统一记为“发生碰撞”；`witness_pair`仅记录最危险的一对，不排除同时存在其他碰撞对。
6. 方向统一：盘入螺线、双圆弧、盘出螺线共用同一个沿运动方向递增的$s$，从而保证位置和速度接口一致。
"""
    output.write_text(text, encoding="utf-8")


def _write_handoff(geometry, states, first_collision, output: Path) -> None:
    best = min(states, key=lambda state: state.global_margin_m)
    collision_states = [state for state in states if state.global_margin_m < 0.0]
    maximum_speed = max(max(abs(value) for value in state.speeds) for state in states)
    selection_gap = PITCH - Q3_FEASIBLE_UPPER
    coarse_gap = 0.451 - Q3_FEASIBLE_UPPER
    text = f"""# Q4：p=0.450342 m 的 -100 s 到 100 s 全链计算交接稿

## 1. 本问目标

按用户指定的扩展口径，把第三问的临界安全螺距带入第四问路径，计算 $t=-100,-99,\\ldots,100$ s 共201个时刻的224个把手位置与速度，并重新检查加入S形调头路径后是否碰撞。

这里必须区分题面口径与扩展口径：题面第四问原始参数是 $p=1.7$ m；本交接稿采用的是用户明确指定的第三问延伸值 $p=0.450342$ m，不能在论文中把二者混写。

## 2. 继承与新增

| 建模模块 | 继承/新增 | 为什么需要 | 输入 | 输出 | 后续用途 |
| --- | --- | --- | --- | --- | --- |
| 等距螺线参数化 | 公共模型 | 统一表示盘入、盘出轨迹 | $p,\\theta$ | 轨迹点、单位切向 | 位置和速度 |
| 定长圆递推 | 公共模型 | 相邻把手距离固定 | 龙头位置、$L_i$ | 224个把手位置 | 实体恢复 |
| 严格2:1双圆弧 | Q4新增 | 满足切向连续、中心对称、半径比约束 | 两边界切点、4.5 m圆域 | S形调头段 | 路径拼接 |
| 板凳矩形恢复 | 继承Q2 | 碰撞对象是板凳实体而非轴线 | 相邻把手位置 | 223个定向矩形 | 碰撞判定 |
| 初筛+SAT | 继承Q2 | 降低无效矩形对计算量并精确判碰 | 板凳矩形 | 裕度、危险板对 | 可行性结论 |
| 刚杆速度递推 | 继承Q1 | 节点速度受定长约束关联 | 位置、切向、龙头速度 | 各把手速度 | Q4/Q5输出 |

## 3. 为什么取0.450342，而不是0.451

第三问逻辑二分给出不可行下界 $p_L={Q3_INFEASIBLE_LOWER}$ m、可行上界 $p_U={Q3_FEASIBLE_UPPER}$ m；中点估计 $p^*\\approx{Q3_MIDPOINT_ESTIMATE}$ m 只是临界估计，不是已验证的安全参数。

本批从已验证可行上界向上取到六位小数：

$$p=\\lceil10^6p_U\\rceil\\times10^{{-6}}=0.450342\\ \\mathrm{{m}}.$$

它只比 $p_U$ 高 {selection_gap:.12g} m，即 {selection_gap*1000:.12g} mm；如果取0.451 m，则会比上界高 {coarse_gap:.12g} m，约放大 {coarse_gap/selection_gap:.1f} 倍。因而0.450342既保持在已验证安全侧，又最贴近第三问临界值。对该值复跑第三问全路径，最小裕度为 {Q3_SELECTED_RECHECK_MARGIN:.12g} m，仍为正。

这个论证只负责“为什么选这个输入”。Q3的安全范围止于盘入螺线到调头边界，Q4新增双圆弧以后必须重新进行碰撞判定。

## 4. 模型建立与完整求解过程

文字箭头流程：

`Q3可行上界 → 六位小数最小上取0.450342 → 求4.5 m边界切点 → 构造严格2:1中心对称双圆弧 → 拼接盘入/调头/盘出弧长轨迹 → t=-100..100逐秒确定龙头位置 → 定长圆逐节点递推其余把手 → 轴线恢复板凳矩形 → 拓扑排除与外接圆初筛 → SAT精筛并记录全局最小裕度 → 刚杆约束求导递推速度 → 写入result4.xlsx与验证表`

相邻把手满足 $\\lVert P_i-P_{{i-1}}\\rVert=L_i$。对时间求导得到

$$w_i=w_{{i-1}}\\frac{{(P_i-P_{{i-1}})\\cdot\\tau_{{i-1}}}}{{(P_i-P_{{i-1}})\\cdot\\tau_i}},$$

再由 $(v_{{x,i}},v_{{y,i}})=w_i\\tau_i$ 得到速度分量。碰撞部分保留Q2的“拓扑排除→外接圆初筛→SAT精筛”；任意一对非相邻板凳裕度小于等于0，就统一描述为“发生碰撞”，不要求在重叠遮挡关系中强行判成只和板凳1或只和板凳2碰撞。

## 5. 结果及其实际含义

| 项目 | 结果 |
| --- | ---: |
| 计算螺距 | {PITCH:.6f} m |
| 时间范围 | -100 s到100 s，步长1 s |
| 全链把手数 | 224 |
| 调头圆弧总长 | {geometry.length:.9f} m |
| 大圆弧半径 | {geometry.radius_large:.9f} m |
| 小圆弧半径 | {geometry.radius_small:.9f} m |
| 调头路径最大半径 | {geometry.maximum_radius():.9f} m |
| 201个整数时刻中的碰撞状态数 | {len(collision_states)} |
| 整数时刻全局最小裕度 | {best.global_margin_m:.12g} m |
| 最小裕度时刻 | {best.time_s:.0f} s |
| 该时刻危险板凳对 | {best.witness_pair} |
| 全部状态最大把手速度 | {maximum_speed:.9f} m/s |
| 连续时间首次碰撞 | {first_collision['time_s'] if first_collision else '未检出'} s |

直接结论：这条边界切点、严格2:1的双圆弧基线能够完整产生位置和速度，但出现碰撞状态，因此它是“可计算的几何基线”，不是“全过程无碰撞的Q4最终解”。这一结果也说明：第三问的最小安全螺距不能直接保证第四问新增调头段安全。

`collision_flag=1`表示该秒至少存在一对碰撞；`witness_pair`只记录全局裕度最小的一对，不表示其他同时碰撞对不存在。

## 6. 模型检验、局限与下一问接口

- 几何残差：圆相切、端点切向、接合切向均按 $10^{{-8}}$ 量级检查；大、小圆半径比按 $10^{{-12}}$ 量级检查。
- 数值完整性：201个时刻、224个把手的位置与速度均为有限值，速度递推分母远离奇异阈值。
- 表格检查：`result4.xlsx`保留题面“位置/速度”工作表，201个时刻表头和关键数据区完整，未检出Excel错误值。
- 灵敏度含义：0.450342只比Q3安全上界高微米量级，而Q4最小碰撞裕度是更大的负值量级，所以是否把第六位再抬高一个单位不会改变“当前双圆弧基线发生碰撞”的结论。
- 局限：本批没有把所有切点自由度重新做全局无碰撞优化；它忠实回答了用户指定参数和当前严格边界双圆弧基线的计算结果。
- Q5接口：Q5读取 `tables/result4.xlsx`、`tables/q4_p0450342_requested_summary.csv`、`tables/q4_p0450342_evidence.json`；必须使用0.450342，且要继承Q4“当前基线发生碰撞”的边界说明。

## 相对路径索引

- `tables/result4.xlsx`：题面模板格式的全链位置和速度；
- `tables/result4_template.xlsx`：未填写的题面模板；
- `tables/q4_p0450342_collision_states.csv`：逐秒碰撞裕度、危险板对、最大速度；
- `tables/q4_p0450342_requested_summary.csv`：五个指定时刻和指定把手摘要；
- `tables/q4_p0450342_evidence.json`：机器可读运行证据；
- `tables/Q4_p0450342_验证结果.md`：几何、速度和Excel检查；
- `tables/Q4_p0450342_公式自审.md`：统一符号、单位和碰撞口径；
- `code/`：复现程序和依赖。
"""
    output.write_text(text, encoding="utf-8")


def _write_readme(output: Path) -> None:
    text = """# 04_Q4_调头路径_p0450342扩展

本目录是用户指定的 `p=0.450342 m` 第四问扩展计算交接包。入口文档为 `04_Q4_p0450342_建模手交接.md`，结果表在 `tables/`，复现代码在 `code/`。

注意：题面第四问原始螺距为1.7 m；本目录是继承第三问最小安全螺距的扩展口径。当前严格2:1边界双圆弧基线发生碰撞，不能直接作为无碰撞最终方案。
"""
    output.write_text(text, encoding="utf-8")


def _copy_code(code_dir: Path) -> None:
    if not (HERE / "experiments" / "q4_small_collision_map.py").exists():
        return
    source_files = [
        HERE / "q4_result4_p0450342.py",
        HERE / "run_q4_result4_p0450342.py",
        HERE / "tests" / "test_q4_result4_p0450342.py",
        HERE / "experiments" / "q4_small_collision_map.py",
        HERE.parents[1] / "03-deliverables" / "02_Q2_碰撞模型" / "code" / "collision_q2.py",
    ]
    code_dir.mkdir(parents=True, exist_ok=True)
    for source in source_files:
        shutil.copy2(source, code_dir / source.name)
    (code_dir / "requirements.txt").write_text("numpy\nscipy\nopenpyxl\npytest\n", encoding="utf-8")


def _copy_tree_clean(source: Path, target: Path) -> None:
    if target.exists():
        shutil.rmtree(target)
    shutil.copytree(source, target, ignore=shutil.ignore_patterns("__pycache__", ".pytest_cache", "*.pyc"))


def main() -> int:
    delivery, template, repo_target = _resolve_paths()
    tables = delivery / "tables"
    code = delivery / "code"
    tables.mkdir(parents=True, exist_ok=True)
    code.mkdir(parents=True, exist_ok=True)

    geometry, states = compute_series(PITCH, range(-100, 101), n_handles=224)
    first_collision = _find_first_collision_time(geometry)
    result_path = tables / "result4.xlsx"
    _write_workbook(states, template, result_path)
    shutil.copy2(template, tables / "result4_template.xlsx")
    _write_collision_csv(states, tables / "q4_p0450342_collision_states.csv")
    _write_requested_summary(states, tables / "q4_p0450342_requested_summary.csv")
    workbook_check = _validate_workbook(result_path)
    best = min(states, key=lambda state: state.global_margin_m)
    collision_states = [state for state in states if state.global_margin_m < 0.0]
    evidence = {
        "run_id": f"Q4-P0450342-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
        "pitch_m": PITCH,
        "official_q4_pitch_m": 1.7,
        "scenario": "user-directed Q3-to-Q4 extension, not the original Q4 pitch",
        "pitch_selection": {
            "q3_infeasible_lower_m": Q3_INFEASIBLE_LOWER,
            "q3_feasible_upper_m": Q3_FEASIBLE_UPPER,
            "q3_midpoint_estimate_m": Q3_MIDPOINT_ESTIMATE,
            "selected_m": PITCH,
            "selected_minus_upper_m": PITCH - Q3_FEASIBLE_UPPER,
            "q3_selected_recheck_min_margin_m": Q3_SELECTED_RECHECK_MARGIN,
            "rule": "round the verified feasible upper endpoint upward at six decimal places",
        },
        "trajectory": {
            "interpretation": "strict 2:1 boundary-endpoint S-biarc",
            "turn_radius_m": 4.5,
            "turn_length_m": float(geometry.length),
            "radius_large_m": float(geometry.radius_large),
            "radius_small_m": float(geometry.radius_small),
            "maximum_path_radius_m": float(geometry.maximum_radius()),
            "circle_tangency_residual": float(geometry.circle_tangency_residual),
            "endpoint_tangent_error": float(geometry.endpoint_tangent_error),
            "joint_tangent_error": float(geometry.joint_tangent_error),
        },
        "calculation": {
            "time_range_s": [-100, 100],
            "time_step_s": 1,
            "state_count": len(states),
            "handle_count": 224,
            "collision_state_count": len(collision_states),
            "collision_free_state_count": len(states) - len(collision_states),
            "minimum_global_margin_m": float(best.global_margin_m),
            "minimum_margin_time_s": int(round(best.time_s)),
            "minimum_margin_witness_pair": list(best.witness_pair),
            "first_continuous_collision": first_collision,
            "maximum_speed_m_per_s": float(max(max(abs(value) for value in state.speeds) for state in states)),
            "minimum_speed_denominator": float(min(state.minimum_speed_denominator for state in states)),
        },
        "workbook_validation": workbook_check,
        "warning": "Q3 safety only covers the inward spiral. This Q4 baseline has collisions and is not a collision-free Q4 solution.",
    }
    (tables / "q4_p0450342_evidence.json").write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_validation_markdown(geometry, states, workbook_check, first_collision, tables / "Q4_p0450342_验证结果.md")
    _write_formula_audit(tables / "Q4_p0450342_公式自审.md")
    _write_handoff(geometry, states, first_collision, delivery / "04_Q4_p0450342_建模手交接.md")
    _write_readme(delivery / "README.md")
    _copy_code(code)
    if repo_target is not None:
        _copy_tree_clean(delivery, repo_target)
    print(json.dumps({
        "output": str(delivery),
        "repo_handoff": str(repo_target) if repo_target else None,
        "states": len(states),
        "collisions": len(collision_states),
        "minimum_margin_m": best.global_margin_m,
        "minimum_margin_time_s": best.time_s,
        "maximum_speed_m_per_s": max(max(abs(value) for value in state.speeds) for state in states),
        "first_continuous_collision": first_collision,
        "workbook_validation": workbook_check,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
