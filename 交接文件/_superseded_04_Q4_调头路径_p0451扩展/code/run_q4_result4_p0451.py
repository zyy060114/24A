"""Run and package the Q4 p=0.451 m extension time series."""

from __future__ import annotations

import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[2]
sys.path.insert(0, str(HERE))
from q4_result4_p0451 import PITCH, build_boundary_geometry, compute_series  # noqa: E402


LOCAL_DELIVERY = ROOT / "03-deliverables" / "04_Q4_调头路径_p0451扩展"
LOCAL_TABLES = LOCAL_DELIVERY / "tables"
LOCAL_CODE = LOCAL_DELIVERY / "code"
REPO_HANDOFF = Path(r"D:\github\24A\交接文件\建模手交接\04_Q4_调头路径_p0451扩展")
TEMPLATE = ROOT / "附件" / "result4.xlsx"


def _write_workbook(states, output: Path) -> None:
    workbook = load_workbook(TEMPLATE)
    position_sheet = workbook["位置"]
    speed_sheet = workbook["速度"]
    times = [int(round(state.time_s)) for state in states]
    for column, time_s in enumerate(times, start=2):
        position_sheet.cell(1, column).value = f"{time_s} s"
        speed_sheet.cell(1, column).value = f"{time_s} s"
    for column, state in enumerate(states, start=2):
        for handle_index, (x, y) in enumerate(state.positions):
            if handle_index <= 221:
                row_x = 2 + 2 * handle_index
            elif handle_index == 222:
                row_x = 446
            else:
                row_x = 448
            position_sheet.cell(row_x, column).value = round(float(x), 6)
            position_sheet.cell(row_x + 1, column).value = round(float(y), 6)
        for handle_index, speed in enumerate(state.speeds):
            speed_sheet.cell(2 + handle_index, column).value = round(abs(float(speed)), 6)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def _write_csv(states, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = ["time_s", "global_margin_m", "collision_flag", "witness_pair", "minimum_speed_denominator", "max_handle_speed_m_per_s"]
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for state in states:
            writer.writerow({
                "time_s": int(round(state.time_s)),
                "global_margin_m": f"{state.global_margin_m:.12g}",
                "collision_flag": int(state.global_margin_m < 0.0),
                "witness_pair": f"{state.witness_pair[0]}-{state.witness_pair[1]}",
                "minimum_speed_denominator": f"{state.minimum_speed_denominator:.12g}",
                "max_handle_speed_m_per_s": f"{max(abs(x) for x in state.speeds):.12g}",
            })


def _write_requested_summary(states, output: Path) -> None:
    wanted_times = {-100, -50, 0, 50, 100}
    wanted_handles = [0, 1, 51, 101, 151, 201, 223]
    fields = ["time_s", "handle_index", "paper_label", "x_m", "y_m", "vx_m_per_s", "vy_m_per_s", "speed_m_per_s"]
    labels = {0: "龙头前把手", 1: "龙头后第1节", 51: "龙头后第51节", 101: "龙头后第101节", 151: "龙头后第151节", 201: "龙头后第201节", 223: "龙尾后把手"}
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for state in states:
            if int(round(state.time_s)) not in wanted_times:
                continue
            for index in wanted_handles:
                x, y = state.positions[index]
                vx, vy = state.velocities[index]
                writer.writerow({
                    "time_s": int(round(state.time_s)),
                    "handle_index": index,
                    "paper_label": labels[index],
                    "x_m": f"{x:.9f}", "y_m": f"{y:.9f}",
                    "vx_m_per_s": f"{vx:.9f}", "vy_m_per_s": f"{vy:.9f}",
                    "speed_m_per_s": f"{abs(state.speeds[index]):.9f}",
                })


def _write_evidence(geometry, states, output: Path) -> None:
    collision_states = [s for s in states if s.global_margin_m < 0.0]
    best = min(states, key=lambda s: s.global_margin_m)
    payload = {
        "run_id": f"Q4-P0451-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
        "pitch_m": PITCH,
        "trajectory_interpretation": "strict 2:1 boundary-endpoint biarc; Q3-derived extension scenario",
        "time_range_s": [-100, 100],
        "time_step_s": 1,
        "handle_count": 224,
        "turn_radius_m": 4.5,
        "turn_length_m": float(geometry.length),
        "radius_large_m": float(geometry.radius_large),
        "radius_small_m": float(geometry.radius_small),
        "maximum_path_radius_m": float(geometry.maximum_radius()),
        "minimum_global_margin_m": float(best.global_margin_m),
        "minimum_margin_time_s": int(round(best.time_s)),
        "minimum_margin_witness_pair": list(best.witness_pair),
        "collision_state_count": len(collision_states),
        "collision_free_state_count": len(states) - len(collision_states),
        "speed_max_m_per_s": float(max(max(abs(x) for x in s.speeds) for s in states)),
        "minimum_speed_denominator": float(min(s.minimum_speed_denominator for s in states)),
        "parameter_reason": {
            "q3_infeasible_lower_m": 0.45033203125,
            "q3_feasible_upper_m": 0.450341796875,
            "q3_midpoint_estimate_m": 0.4503369140625,
            "selected_m": 0.451,
            "explanation": "The upper endpoint is the first tested full-path collision-free bound; 0.451 m is its upward millimetre-level conservative report value. The midpoint is only a threshold estimate and is not used as a guaranteed-safe input."
        },
        "warning": "The Q3 safety result does not automatically transfer to the Q4 turn. This p=0.451 boundary-biarc extension has collision states and is not a collision-free Q4 solution.",
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_handoff(geometry, states, output: Path) -> None:
    best = min(states, key=lambda s: s.global_margin_m)
    collision_states = [s for s in states if s.global_margin_m < 0.0]
    text = f"""# Q4：p=0.451 m 的 -100 s 到 100 s 全链计算交接稿

## 1. 本批计算口径

本批按用户指定，将第三问的保守螺距取值 **p=0.451 m** 代入第四问扩展计算。这里必须明确：题面第四问原始参数写的是 p=1.7 m，因此本文件不是原题第四问的唯一官方口径，而是“继承第三问最小安全螺距的扩展口径”。

调头轨迹采用调头圆边界两端点、严格半径比 $R_1:R_2=2:1$ 的双圆弧基线；圆弧与两条螺线相切，圆弧最大半径约束为 4.5 m。时间取整数 $t=-100,-99,\ldots,100$ s，共201个状态；每个状态恢复224个把手的位置，并依据刚杆微分约束递推速度。

## 2. 为什么取 0.451 m

第三问全过程碰撞计算得到：

- $p=0.45033203125$ m：最小裕度为负，属于不可行下界；
- $p=0.450341796875$ m：最小裕度为正，属于可行上界；
- $p^*\approx0.4503369140625$ m：只是上下界中点的临界估计，不是保证安全的参数。

因此本批不取临界中点，而取已验证可行上界并向毫米级向上报告：

$$p=0.451\,\mathrm{{m}}.$$

这个选择的逻辑是“下界碰撞、上界不碰撞、中点只估计、向上取整留出数值和报告裕度”。但该安全性只对第三问盘入螺线全过程成立，不能自动外推到第四问加入双圆弧后的新轨迹，所以本批仍重新检查 Q4 碰撞。

## 3. 直接结果

| 项目 | 结果 |
| --- | ---: |
| 计算螺距 | {PITCH:.3f} m |
| 时间范围 | -100 s 到 100 s，步长1 s |
| 全链把手数 | 224 |
| 调头圆弧总长 | {geometry.length:.9f} m |
| 大圆弧半径 | {geometry.radius_large:.9f} m |
| 小圆弧半径 | {geometry.radius_small:.9f} m |
| 调头路径最大半径 | {geometry.maximum_radius():.9f} m |
| 全部201个状态中碰撞状态数 | {len(collision_states)} |
| 全局最小裕度 | {best.global_margin_m:.12g} m |
| 最小裕度时刻 | {best.time_s:.0f} s |
| 危险板凳对 | {best.witness_pair} |
| 全部状态最大把手速度 | {max(max(abs(x) for x in s.speeds) for s in states):.9f} m/s |

**结论：这条 p=0.451 m 的边界双圆弧基线可以完整计算位置和速度，但在加入调头路径后出现碰撞状态，因此不能写成“Q4 已找到全过程无碰撞路径”。**

## 4. 文字箭头流程图

`第三问可行上界 → 向毫米级保守取 p=0.451 → 构造4.5 m边界严格2:1双圆弧 → 拼接盘入/调头/盘出分段弧长轨迹 → t=-100..100逐秒定位龙头 → 定长圆递推223个后续把手 → 恢复矩形板凳 → SAT计算全局碰撞裕度 → 刚杆微分约束递推速度 → 写入result4.xlsx与验证表`

## 5. 文件位置

- `tables/result4.xlsx`：题面模板格式的全链位置和速度；
- `tables/q4_p0451_collision_states.csv`：每秒碰撞裕度、危险板对和最大速度；
- `tables/q4_p0451_requested_summary.csv`：论文要求的五个时刻和指定把手；
- `tables/q4_p0451_evidence.json`：机器可读运行证据；
- `code/q4_result4_p0451.py`、`code/run_q4_result4_p0451.py`：复现代码。

详细结果回到总控汇报；本批不启动第五问。
"""
    output.write_text(text, encoding="utf-8")


def main() -> int:
    geometry, states = compute_series(PITCH, range(-100, 101), n_handles=224)
    LOCAL_TABLES.mkdir(parents=True, exist_ok=True)
    LOCAL_CODE.mkdir(parents=True, exist_ok=True)
    _write_workbook(states, LOCAL_TABLES / "result4.xlsx")
    _write_csv(states, LOCAL_TABLES / "q4_p0451_collision_states.csv")
    _write_requested_summary(states, LOCAL_TABLES / "q4_p0451_requested_summary.csv")
    _write_evidence(geometry, states, LOCAL_TABLES / "q4_p0451_evidence.json")
    _write_handoff(geometry, states, LOCAL_DELIVERY / "04_Q4_p0451_建模手交接.md")
    shutil.copy2(HERE / "q4_result4_p0451.py", LOCAL_CODE / "q4_result4_p0451.py")
    shutil.copy2(HERE / "run_q4_result4_p0451.py", LOCAL_CODE / "run_q4_result4_p0451.py")
    REPO_HANDOFF.mkdir(parents=True, exist_ok=True)
    if REPO_HANDOFF.exists():
        for source in [LOCAL_DELIVERY / "04_Q4_p0451_建模手交接.md", LOCAL_TABLES / "result4.xlsx", LOCAL_TABLES / "q4_p0451_collision_states.csv", LOCAL_TABLES / "q4_p0451_requested_summary.csv", LOCAL_CODE / "q4_result4_p0451.py", LOCAL_CODE / "run_q4_result4_p0451.py"]:
            target = REPO_HANDOFF / ("tables" if source.suffix in {".xlsx", ".csv"} else "code" if source.suffix == ".py" else "") / source.name
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
    print(json.dumps({
        "output": str(LOCAL_DELIVERY),
        "repo_handoff": str(REPO_HANDOFF),
        "states": len(states),
        "collisions": sum(s.global_margin_m < 0.0 for s in states),
        "minimum_margin_m": min(s.global_margin_m for s in states),
        "maximum_speed_m_per_s": max(max(abs(x) for x in s.speeds) for s in states),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
